"""
医療費集計フォーム生成

国税庁「医療費集計フォーム」ver3.1 の公式テンプレートにデータを記入して出力する。
マイナポータル連携で処理されない分（保険適用外の控除対象）のみ記載。

テンプレート: data/iryouhi_form_v3.1.xlsx  (NTAからダウンロードした公式フォーム)
出力: output/medical_expense_form.xlsx
"""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
TEMPLATE = BASE / "templates" / "iryouhi_form_v3.1.xlsx"

# テンプレートの列マッピング (row 9 から開始)
DATA_START_ROW = 9
COL_PATIENT = "B"       # 医療を受けた人
COL_FACILITY = "C"      # 病院・薬局などの名称
# D-G: 医療費の区分（チェックボックス式 — "該当する" を入力）
COL_CAT_TREATMENT = "D"  # 診療・治療
COL_CAT_MEDICINE = "E"   # 医薬品購入
COL_CAT_NURSING = "F"    # 介護保険サービス
COL_CAT_OTHER = "G"      # その他の医療費
COL_AMOUNT = "H"         # 支払った医療費の金額
COL_REIMBURSE = "I"      # 左のうち、補填される金額
CHECKBOX_VALUE = "該当する"
MAX_DATA_ROWS = 995


def load_match_results(data_dir: Path) -> dict:
    return json.loads((data_dir / "match_results.json").read_text("utf-8"))


def load_reimbursements(data_dir: Path) -> list[dict]:
    """補填データ (Aflac給付金、自治体補助金等) を読み込む。"""
    reimb_file = data_dir / "reimbursements.json"
    if not reimb_file.exists():
        return []
    return json.loads(reimb_file.read_text("utf-8"))


def apply_reimbursements(rows: list[dict], reimbursements: list[dict]) -> None:
    """補填金額を (patient, facility) が一致する行に加算する。"""
    # (patient, facility) → 補填合計
    reimb_by_key: dict[tuple, int] = defaultdict(int)
    for r in reimbursements:
        reimb_by_key[(r["patient"], r["facility"])] += r["amount"]

    for row in rows:
        key = (row["patient"], row["facility"])
        if key in reimb_by_key:
            row["reimbursement"] += reimb_by_key[key]
            # 補填は対応する医療費を超えない
            row["reimbursement"] = min(row["reimbursement"], row["amount"])
            del reimb_by_key[key]  # 重複適用を防ぐ（最初にマッチした行に適用）

    # 未適用の補填があれば警告
    for key, amount in reimb_by_key.items():
        print(f"警告: 補填 {key} ({amount:,}円) に対応する行が集計フォームにありません")


def collect_xlsx_rows(results: list[dict]) -> list[dict]:
    """集計フォームに記載すべき全項目を収集し、(患者, 医療機関, 区分列) 別にグループ化。

    Returns:
        list of dict with keys: patient, facility, category_col (D/E/F/G), amount, reimbursement
    """

    # (patient, facility, category_col) → 合計金額
    grouped: dict[tuple, int] = defaultdict(int)

    for r in results:
        # 1. include_in_xlsx=True の項目（receipt_only の確定控除分）
        if r.get("include_in_xlsx") and r.get("deductible_amount", 0) > 0:
            patient = r["patient"]
            facility = r["facility"]
            oop = r.get("oop_judgment")
            if oop and oop["deductible_total"] > 0:
                # 保険診療分
                ins_amount = r.get("receipt_insurance_amount", 0)
                if ins_amount > 0:
                    grouped[(patient, facility, COL_CAT_TREATMENT)] += ins_amount
                # OOP控除分
                for item in oop["items"]:
                    if item["judgment"] == "deductible":
                        cat_col = _categorize_item(item["item"])
                        grouped[(patient, facility, cat_col)] += item["amount"]
            else:
                grouped[(patient, facility, COL_CAT_TREATMENT)] += r["deductible_amount"]

        # 2. oop_judgment に deductible 項目がある matched 分
        if not r.get("include_in_xlsx"):
            oop = r.get("oop_judgment")
            if not oop or oop["deductible_total"] == 0:
                continue
            patient = r["patient"]
            facility = r["facility"]
            for item in oop["items"]:
                if item["judgment"] == "deductible":
                    cat_col = _categorize_item(item["item"])
                    grouped[(patient, facility, cat_col)] += item["amount"]

    rows = []
    for (patient, facility, category_col), amount in sorted(grouped.items()):
        rows.append({
            "patient": patient,
            "facility": facility,
            "category_col": category_col,
            "amount": amount,
            "reimbursement": 0,
        })
    return rows


def _categorize_item(item_name: str) -> str:
    """OOP項目名をNTAフォームの区分列 (D/E/F/G) にマッピング。

    D: 診療・治療 — 病院・診療所での診察、不妊治療、出生前診断、妊婦健診等
    E: 医薬品購入 — 処方薬・市販薬
    F: 介護保険サービス
    G: その他の医療費 — 通院交通費、治療用器具 等
    """
    # 医薬品
    if "薬" in item_name or "処方" in item_name:
        return COL_CAT_MEDICINE
    # 介護
    if "介護" in item_name:
        return COL_CAT_NURSING
    # その他: 治療用器具、交通費 等
    if "血圧計" in item_name or "治療用器具" in item_name:
        return COL_CAT_OTHER
    if "交通" in item_name:
        return COL_CAT_OTHER
    # デフォルト: 診療・治療（不妊治療、出生前診断、妊婦健診等すべて含む）
    return COL_CAT_TREATMENT


def create_medical_expense_form(match_data: dict, data_dir: Path, output_dir: Path) -> None:
    """公式テンプレートにデータを記入して保存する。

    テンプレートの構造を一切変更せず、unlocked なデータセル (B-J, row 9+) のみに書き込む。
    合計はテンプレート側の SUM 数式 (C3, C4) が自動計算する。
    """
    if not TEMPLATE.exists():
        raise FileNotFoundError(
            f"公式テンプレートが見つかりません: {TEMPLATE}\n"
            "NTAサイトからダウンロードしてください: "
            "https://www.nta.go.jp/taxes/shiraberu/shinkoku/tokushu/keisubetsu/iryou-shuukei.htm"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "medical_expense_form.xlsx"

    results = match_data["results"]
    rows = collect_xlsx_rows(results)

    # 補填データを適用
    reimbursements = load_reimbursements(data_dir)
    if reimbursements:
        apply_reimbursements(rows, reimbursements)

    if len(rows) > MAX_DATA_ROWS:
        print(f"警告: データ行数 {len(rows)} が上限 {MAX_DATA_ROWS} を超えています。先頭 {MAX_DATA_ROWS} 行のみ記載します。")
        rows = rows[:MAX_DATA_ROWS]

    # テンプレートをコピーして開く
    shutil.copy2(TEMPLATE, out_path)
    wb = load_workbook(out_path)
    ws = wb["医療費集計フォーム"]

    # シート保護を一時解除してデータを書き込み
    ws.protection.sheet = False

    for i, item in enumerate(rows):
        row_num = DATA_START_ROW + i
        ws[f"{COL_PATIENT}{row_num}"] = item["patient"]
        ws[f"{COL_FACILITY}{row_num}"] = item["facility"]
        # 医療費の区分: 該当する列にチェック値を入力
        ws[f"{item['category_col']}{row_num}"] = CHECKBOX_VALUE
        ws[f"{COL_AMOUNT}{row_num}"] = item["amount"]
        if item["reimbursement"]:
            ws[f"{COL_REIMBURSE}{row_num}"] = item["reimbursement"]

    # シート保護を再有効化
    ws.protection.sheet = True

    wb.save(out_path)

    total = sum(r["amount"] for r in rows)
    total_reimb = sum(r["reimbursement"] for r in rows)
    print(f"医療費集計フォームを {out_path} に出力しました")
    print(f"  記載件数: {len(rows)}行")
    print(f"  支払った医療費合計: {total:,}円")
    print(f"  補てんされる金額合計: {total_reimb:,}円")


def main(data_dir: Path, output_dir: Path):
    match_data = load_match_results(data_dir)
    create_medical_expense_form(match_data, data_dir, output_dir)


def run(year_dir: Path, year_label: str) -> None:
    """パイプラインから呼ばれるエントリポイント。"""
    data_dir = year_dir / "data"
    output_dir = year_dir / "output"
    main(data_dir, output_dir)


if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 2:
        year_dir = BASE / "years" / sys.argv[1]
        run(year_dir, sys.argv[1])
    else:
        # 後方互換: 旧パスで実行
        main(BASE / "data", BASE / "output")
